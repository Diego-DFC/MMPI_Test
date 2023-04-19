# -*- coding:utf-8 -*-

# 版权所有 (C) 2018.6.25 金盛羽。保留所有权利。
# Copyright 2018.6.25 Shengyu Jin. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
本模块主用于处理和MMPI测试相关的函数功能，其中包含一份完整的MMPI测试问卷
This module is mainly for processing functions about MMPI, including
a complete questionnaire.
"""

import time
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Font, Alignment

# for debug
from random import randint

# 指导语
# Instruction
Ins1 = """-----------------------------------------------------------------
                       明尼苏达多项人格测验                      
       Minnesota Multiphasic Per-sonality Inventory (MMPI)       
-----------------------------------------------------------------
Breve descripción: el Inventario de Personalidad Múltiple de Minnesota (MMPI) fue desarrollado por S. R. Hathaway y J. C. McKinley, profesores de la Universidad de Minnesota, en 1942.
（S. R. Hathaway y J. C. McKinley en 1942.
Puede utilizarse para comprobar el tipo de personalidad de una persona normal o para distinguir entre una persona normal y una persona con una enfermedad mental o psiquiátrica. A finales del siglo pasado
A finales del siglo pasado, sobre la base de muchos años de investigación y la investigación práctica por el Sr. Ji Jumao y otros, el cuestionario MMPI y el Inventario Psicológico Frecuente (IPF) fueron desarrollados para su uso en China.
El cuestionario y la norma del MMPI se utilizan en China desde 1980 y se han extendido su uso en los últimos años.
Los resultados demuestran que la prueba tiene un cierto grado de fiabilidad y validez en China y posee un alto valor clínico de referencia.
-----------------------------------------------------------------
Recordatorio importante: Este test es un instrumento psicométrico profesional con valor de referencia clínico y debe ser utilizado e interpretado con precaución por los profesionales pertinentes.
Debe ser utilizado e interpretado bajo la orientación de un profesional pertinente, ¡y debe ser utilizado con precaución por particulares!                                 
-----------------------------------------------------------------"""

Ins2 = """-----------------------------------------------------------------
Orientación: Este cuestionario consta de una serie de preguntas relevantes para usted.
si coincide con su comportamiento, sentimientos, actitudes u opiniones actuales. En caso afirmativo, marque un "1"; en caso contrario, marque un "0".
Escriba un "0" en caso afirmativo, en caso contrario escriba un "0" y confirme pulsando la tecla "Intro" para completar la pregunta. Escriba sus primeras impresiones lo antes posible tras leer la pregunta.
No dedique demasiado tiempo a reflexionar sobre cada pregunta. Las personalidades varían y no existe una respuesta correcta o incorrecta.
No hay que preocuparse por lo que es correcto o incorrecto, bueno o malo, simplemente responda como mejor le parezca.
-----------------------------------------------------------------
Ejemplo de respuesta:
x. Mi sexo es
1. masculino 0. femenino
> 1
-----------------------------------------------------------------"""


# Preguntas del cuestionario MMPI de 566 ítems
# 566 questions of MMPI questionnaire
Que = {
    1: 'Me gustan las revistas de mecánica.',
    2: 'Tengo buen apetito.',
    3: 'Casi siempre me levanto por las mañanas descansado y como nuevo.',
    4: 'Creo que me gustaría el trabajo de bibliotecario(a).',
    5: 'El ruido me despierta fácilmente.',
    6: 'Mi padre es un buen hombre o (si ya ha fallecido) mi padre era un buen hombre.',
    7: 'Me gusta leer artículos sobre crímenes en los periódicos.',
    8: 'Normalmente tengo bastante calientes los pies y las manos.',
    9: 'En mi vida diaria hay muchas cosas que me resultan interesantes.',
    10: 'Actualmente tengo tanta capacidad de trabajo como antes.',
    11: 'Muy a menudo me parece que tengo un nudo en la garganta.',
    12: 'Una persona debiera tratar de comprender sus sueños, guiarse por ellos o tenerlos en cuenta como avisos.',
    13: 'Me gustan los cuentos de detectives o de misterio.',
    14: 'Trabajo bajo una tensión muy grande.',
    15: 'Tengo diarrea una vez al mes o más frecuentemente.',
    16: 'De vez en cuando pienso en cosas demasiado malas como para hablar de ellas.',
    17: 'Estoy seguro de que la vida me trata mal.',
    18: 'Muy rara vez sufro de estitiquez.',
    19: 'Cuando acepto un nuevo trabajo me gusta que me indiquen a quien debo halagar.',
    20: 'Mi vida sexual es satisfactoria.',
    21: 'A veces he sentido un inmenso deseo de abandonar mi hogar.',
    22: 'A veces me dan ataques de risa o de llanto que no puedo controlar.',
    23: 'Sufro de ataques de náuseas y de vómito.',
    24: 'Nadie parece comprenderme.',
    25: 'Me gustaría ser cantante.',
    26: 'Creo que lo mejor es quedarme callado cuando estoy en dificultades.',
    27: 'A veces los malos espíritus se posesionan de mi.',
    28: 'Cuando alguien me hace un mal, siento que debería pagarle con la misma moneda, si es que puedo, por principio.',
    29: 'Padezco de acidez estomacal varias veces por semana.',
    30: 'A veces siento ganas de decir palabrotas (garabatos).',
    31: 'Me dan pesadillas con mucha frecuencia.',
    32: 'Me cuesta concentrarme en una tarea o trabajo.',
    33: 'He tenido experiencias muy peculiares y extrañas.',
    34: 'Casi siempre tengo tos.',
    35: 'Si la gente no la hubiera agarrado conmigo, yo habría tenido mucho más éxito.',
    36: 'Raras veces me preocupo por mi salud.',
    37: 'Nunca me he visto en dificultades a causa de mi conducta sexual.',
    38: 'Durante un tiempo, cuando era chico(a), me robé cosas sin importancia.',
    39: 'A veces siento deseos de romper o quebrar cosas.',
    40: 'Casi siempre preferiría sentarme a soñar despierto a hacer cualquier cosa.',
    41: 'He tenido períodos de días, semanas o meses en que no podía ocuparme de nada porque no tenía el ánimo para hacerlo.',
    42: 'A mi familia no le gusta el trabajo que he escogido (o el que pienso escoger para el resto de mi vida).',
    43: 'Mi sueño es sobresaltado e intranquilo.',
    44: 'La mayor parte del tiempo parece que me duele toda la cabeza.',
    45: 'No siempre digo la verdad.',
    46: 'Mi capacidad de raciocinio es ahora mejor que nunca.',
    47: 'Una vez por semana o más, sin causa aparente, repentinamente siento todo el cuerpo acalorado.',
    48: 'Cuando estoy con gente me perturba el escuchar cosas muy extrañas.',
    49: 'Sería mejor si casi todas las leyes fueran dejadas de lado.',
    50: 'Mi alma a veces abandona mi cuerpo.',
    51: 'Me encuentro en tan buenas condiciones físicas como la mayoría de mis amigos.',
    52: 'Cuando me encuentro en la calle con amigos de colegio o con personas conocidas a quienes no he visto desde hace mucho tiempo, prefiero hacerme el desentendido, a menos que ellos me hablen primero.',
    53: 'Un sacerdote puede curar enfermedades rezando e imponiendo las manos sobre la cabeza del enfermo.',
    54: 'Le agrado a la mayor parte de la gente que me conoce.',
    55: 'Casi nunca he sentido dolores en el corazón o en el pecho.',
    56: 'Cuando chico(a) me suspendieron del colegio una o más veces por hacer la cimarra.',
    57: 'Soy una persona sociable.',
    58: 'Todo está ocurriendo tal como los profetas de la Biblia lo predijeron.',
    59: 'Con frecuencia he tenido que recibir órdenes de alguien que sabía menos que yo.',
    60: 'No leo diariamente todos los editoriales del periódico.',
    61: 'No he vivido la vida correctamente.',
    62: 'Con frecuencia siento como un ardor, punzadas, hormigueo o adormecimiento en algunas partes del cuerpo.',
    63: 'No he tenido dificultad ni para comenzar ni para retener el movimiento intestinal (defecación).',
    64: 'A veces insisto tanto en algo hasta que la gente pierde la paciencia conmigo.',
    65: 'Quise (quiero) a mi padre.',
    66: 'Veo cosas, animales o gente a mi alrededor que otros no ven.',
    67: 'Quisiera ser tan feliz como otras personas parecen ser.',
    68: 'Muy raras veces siento dolor en la nuca.',
    69: 'Me siento fuertemente atraído(a) por personas de mi mismo sexo.',
    70: 'Me gustaba jugar a las prendas.',
    71: 'Creo que mucha gente exagera sus desgracias para ganar la compasión y ayuda de los demás.',
    72: 'Sufro de malestares en la boca del estómago varias veces a la semana o con más frecuencia.',
    73: 'Soy una persona importante.',
    74: 'A menudo he deseado ser mujer. (o si usted es mujer): nunca me ha pesado ser mujer.',
    75: 'Algunas veces me enojo.',
    76: 'Casi siempre me siento melancólico.',
    77: 'Me gusta leer novelas de amor.',
    78: 'Me gusta la poesía.',
    79: 'No me siento herido(a) con facilidad.',
    80: 'De vez en cuando mortifico a los animales.',
    81: 'Creo que me gustaría trabajar como guardabosques.',
    82: 'Es fácil ganarme una discusión.',
    83: 'Cualquier persona capaz y dispuesta a trabajar duro, tiene buenas posibilidades de obtener éxito.',
    84: 'En estos días me es difícil no perder la esperanza de llegar a lograr algo.',
    85: 'A veces me siento tan atraído(a) por efectos personales de otra gente, como zapatos, guantes etc., que me dan ganas de tocarlos o robármelos aunque no me sirvan para nada.',
    86: 'Decididamente no tengo confianza en mí mismo.',
    87: 'Me gustaría ser florista.',
    88: 'Generalmente siento que la vida vale la pena.',
    89: 'Se necesita discutir mucho para convencer la verdad a la mayor parte de la gente.',
    90: 'De vez en cuando dejo para mañana lo que debiera hacer hoy.',
    91: 'No me molesta que se rían de mí.',
    92: 'Me gustaría ser enfermero o enfermera.',
    93: 'Creo que la mayoría de la gente mentiría para salir adelante.',
    94: 'Muchas veces hago cosas de las que me arrepiento después (me arrepiento de más cosas con más frecuencia de lo que otra gente parece arrepentirse).',
    95: 'Voy a la iglesia casi todas las semanas.',
    96: 'Tengo muy pocas peleas con miembros de mi familia.',
    97: 'A veces siento un fuerte impulso de hacer algo dañino o escandaloso.',
    98: 'Creo en la segunda venida de Cristo.',
    99: 'Me gustaría ir a fiestas y a otras reuniones donde haya mucha alegría y ruido.',
    100: 'He encontrado problemas tan llenos de alternativas que me ha sido imposible llegar a una decisión.',
    101: 'Creo que la mujer debe tener tanta libertad sexual como el hombre.',
    102: 'Mis luchas más difíciles son conmigo mismo.',
    103: 'Casi nunca me dan calambres o espasmos musculares.',
    104: 'Me importa poco lo que me pase.',
    105: 'A veces, cuando no me siento bien, estoy mal humorado.',
    106: 'Muchas veces siento como si hubiera hecho algo incorrecto o perverso.',
    107: 'Casi siempre estoy contento.',
    108: 'Parece que mi cabeza o mi nariz están congestionadas la mayor parte del tiempo.',
    109: 'Algunas personas son tan dominantes, que siento el deseo de hacer lo contrario de lo que me piden, aunque sepa que tienen razón.',
    110: 'Alguien me tiene mala voluntad.',
    111: 'Nunca he hecho algo peligroso solo por el gusto de hacerlo.',
    112: 'Con frecuencia siento la necesidad de luchar por lo que creo que es justo.',
    113: 'Creo que la ley debe hacerse cumplir.',
    114: 'A menudo siento como si tuviera un cintillo apretándome la cabeza.',
    115: 'Creo en otra vida después de esta.',
    116: 'Disfruto más de una carrera o en un juego cuando apuesto.',
    117: 'La mayoría de la gente es honrada principalmente por el temor a ser descubierta.',
    118: 'En el colegio a veces me llevaron ante el director por hacer la cimarra.',
    119: 'Mi manera de hablar es como ha sido siempre(ni más ligero, ni más despacio, ni balbuceante, ni ronca).',
    120: 'Mis modales en la mesa no son tan correctos en casa como cuando salgo a comer fuera en compañía de otros.',
    121: 'Creo que están conspirando contra mí.',
    122: 'Me parece que soy tan capaz e inteligente como la mayor parte de los que me rodean.',
    123: 'Creo que me están siguiendo.',
    124: 'La mayor parte de la gente se vale de medios algo ilícitos para obtener beneficios o ventajas antes que perderlos.',
    125: 'Sufro mucho de trastornos estomacales.',
    126: 'Me gustaría ir al teatro.',
    127: 'Sé quien es el responsable de la mayoría de mis problemas.',
    128: 'Ver sangre ni me asusta ni me asquea.',
    129: 'A menudo no puedo comprender por qué he estado tan irritable y malhumorado.',
    130: 'Nunca he vomitado o escupido sangre.',
    131: 'No tengo miedo a contagiarme con enfermedades.',
    132: 'Me gusta recoger flores o cultivar plantas decorativas.',
    133: 'Nunca me he entregado a prácticas sexuales fuera de lo común.',
    134: 'A veces los pensamientos pasan por mi mente con mayor rapidez que lo que puedo expresarlos en palabras.',
    135: 'Si pudiera entrar en un cine sin pagar y estuviera seguro de no ser visto, probablemente lo haría.',
    136: 'Generalmente pienso qué segunda intención puede tener otra persona cuando me hace un favor.',
    137: 'Creo que mi vida es tan agradable como la de la mayor parte de la gente que conozco.',
    138: 'La crítica o los retos me hieren profundamente.',
    139: 'Algunas veces siento el deseo de herirme o de herir a otros.',
    140: 'Me gusta cocinar.',
    141: 'Mi conducta está determinada en gran medida por las costumbres de quienes me rodean.',
    142: 'A veces siento que realmente no sirvo para nada.',
    143: 'Cuando niño(a) pertenecí aun grupo o pandilla que trataba de mantenerse unido contra viento y marea.',
    144: 'Me gustaría ser soldado.',
    145: 'A veces siento el deseo de empezar una pelea a puñetazos con alguien.',
    146: 'Me siento impulsado(a) hacia la vida errante, y nunca me siento feliz a menos que esté viajando de un lado a otro.',
    147: 'Muchas veces he perdido una oportunidad porque no he podido decidirme a tiempo.',
    148: 'Me molesta que me pidan consejo o que me interrumpan cuando estoy trabajando en algo importante.',
    149: 'Acostumbraba a llevar un diario de vida.',
    150: 'Prefiero ganar a perder en un juego.',
    151: 'Alguien ha estado tratando de envenenarme.',
    152: 'Casi todas las noches puedo dormirme sin tener pensamientos o ideas que me preocupen.',
    153: 'Durante los últimos años he gozado de salud la mayor parte del tiempo.',
    154: 'Nunca he tenido un ataque o convulsiones.',
    155: 'No estoy perdiendo ni ganando peso.',
    156: 'He tenido épocas en las que he hecho cosas que luego no recuerdo haber hecho.',
    157: 'Creo que frecuentemente he sido castigado sin motivo.',
    158: 'Lloro con facilidad.',
    159: 'No puedo entender lo que leo, tan bien como lo hacía antes.',
    160: 'Nunca me he sentido mejor que ahora.',
    161: 'A veces siento adolorido la parte superior de la cabeza.',
    162: 'Me molesta que una persona me tome el pelo tan hábilmente que tenga que admitir que me engañaron.',
    163: 'No me canso con facilidad.',
    164: 'Me gusta leer y estudiar acerca de las cosas en que estoy trabajando.',
    165: 'Me gusta conocer gente importante porque eso me hace sentir importante.',
    166: 'Siento miedo cuando miro hacia abajo desde un lugar alto.',
    167: 'No me sentiría nervioso si algún familiar mío tuviera dificultades con la ley.',
    168: 'Hay algo que no anda bien en mi mente.',
    169: 'No tengo miedo a manejar dinero.',
    170: 'No me preocupa lo que otros piensen de mí.',
    171: 'Me siento incómodo cuando tengo que hacer una payasada en una reunión, aún cuando otros estén haciendo lo mismo.',
    172: 'Frecuentemente tengo que esforzarme para demostrar que no soy tímido.',
    173: 'Me gustaba el colegio.',
    174: 'Nunca me he desmayado.',
    175: 'Rara vez o nunca he tenido mareos.',
    176: 'No le tengo mucho miedo a las serpientes.',
    177: 'Mi madre era (es) una buena mujer.',
    178: 'Mi memoria parece ser buena.',
    179: 'Me preocupan las cuestiones sexuales.',
    180: 'Me cuesta entablar conversación con gente que recién conozco.',
    181: 'Cuando me siento aburrido me gusta promover algo emocionante.',
    182: 'Tengo miedo de volverme loco(a).',
    183: 'Estoy en contra de dar dinero a los mendigos.',
    184: 'Frecuentemente oigo voces sin saber de donde vienen.',
    185: 'Aparentemente oigo tan bien como la mayoría de las personas.',
    186: 'Con frecuencia noto que mis manos tiemblan cuando trato de hacer algo.',
    187: 'Las manos no se me han puesto torpes o poco hábiles.',
    188: 'Puedo leer por largo rato sin que se me cansen los ojos.',
    189: 'Siento debilidad general la mayor parte del tiempo.',
    190: 'Muy pocas veces me duele la cabeza.',
    191: 'Algunas veces, cuando paso una vergüenza, empiezo a transpirar, lo que me molesta muchísimo.',
    192: 'No he tenido dificultades en mantener el equilibrio cuando camino.',
    193: 'No me dan ataques de alergia o asma.',
    194: 'He tenido ataques durante los cuales no podía controlar mis movimientos o el habla, pero me daba cuenta de lo que ocurría a mi alrededor.',
    195: 'No me agradan todas las personas que conozco.',
    196: 'Me gusta visitar lugares donde nunca he estado.',
    197: 'Alguien ha tratado de robarme.',
    198: 'Muy pocas veces sueño despierto.',
    199: 'Se debe enseñar a los niños los hechos fundamentales del sexo.',
    200: 'Hay personas que están tratando de apoderarse de mis pensamientos o ideas.',
    201: 'Desearía no ser tan tímido.',
    202: 'Creo que estoy condenado o no tengo salvación.',
    203: 'Si yo fuera reportero me gustaría escribir noticias de teatro.',
    204: 'Me gustaría ser periodista.',
    205: 'A veces me ha sido imposible evitar el robar o llevarme algo de una tienda.',
    206: 'Soy muy religioso (a) (más que la mayoría de la gente).',
    207: 'Me gustan distintas clases de juegos o diversiones.',
    208: 'Me gusta coquetear.',
    209: 'Creo que mis pecados son imperdonables.',
    210: 'Todo tiene el mismo sabor.',
    211: 'Puedo dormir de día pero no de noche.',
    212: 'Mi familia me trata más como niño que como adulto.',
    213: 'Cuando camino tengo mucho cuidado de no pisar las líneas de las veredas.',
    214: 'Nunca he tenido erupciones en la piel que me hayan preocupado.',
    215: 'He bebido alcohol en exceso.',
    216: 'Hay muy poco compañerismo y cariño en mi familia en comparación con otros hogares.',
    217: 'Frecuentemente me encuentro preocupado por algo.',
    218: 'No me molesta especialmente el ver sufrir a los animales.',
    219: 'Creo que me gustaría el trabajo de contratista de obras.',
    220: 'Yo quise (quiero) a mi madre.',
    221: 'Me gusta la ciencia.',
    222: 'No me cuesta pedir ayuda a mis amigos aunque no pueda devolverles el favor.',
    223: 'Me gusta mucho cazar.',
    224: 'Con frecuencia mis padres han objetado el tipo de gente con la que acostumbraba a andar.',
    225: 'A veces chismorreo un poco (digo pelambres).',
    226: 'Algunos de mis familiares tienen hábitos que me molestan y me dan mucha rabia.',
    227: 'Me han dicho que camino dormido.',
    228: 'A veces siento que puedo decidirme con extraordinaria serenidad.',
    229: 'Me gustaría pertenecer a varios clubes o asociaciones.',
    230: 'Rara vez noto los latidos de mi corazón y muy pocas veces me falta el aliento.',
    231: 'Me gusta hablar sobre temas sexuales.',
    232: 'He sido criado (formado) en un modo de vida basado en el deber, y lo he seguido siempre fielmente.',
    233: 'Algunas veces he sido un obstáculo a personas que querían hacer algo, no porque eso fuera de mucha importancia, sino por cuestión de principio.',
    234: 'Me enojo con facilidad pero se me pasa pronto.',
    235: 'He sido bastante dependiente y libre de la disciplina familiar.',
    236: 'Medito mucho las cosas, o le doy muchas vueltas a las cosas.',
    237: 'Casi todos mis parientes me apoyan.',
    238: 'Tengo períodos de tanta intranquilidad que no puedo permanecer sentado en una silla por mucho tiempo.',
    239: 'He sufrido desengaños amorosos.',
    240: 'Nunca me preocupo por mi aspecto.',
    241: 'Sueño frecuentemente acerca de cosas que es mejor mantenerlas en secreto.',
    242: 'Creo que no soy más nervioso que la mayoría de las personas.',
    243: 'Sufro de pocos o ninguna clase de dolor.',
    244: 'Mi modo de hacer las cosas tiende a ser mal interpretado por otros.',
    245: 'Mis padres y familiares me encuentran más defectos de los que debieran.',
    246: 'Con frecuencia me salen manchas rojas en el cuello.',
    247: 'Tengo motivos para sentirme celoso de uno o más miembros de mi familia.',
    248: 'A veces, sin motivo, aunque las cosas no me estén saliendo bien, me siento muy alegre, “como un rey”.',
    249: 'Creo que existen el diablo y el infierno en la otra vida.',
    250: 'No culpo a nadie por tratar de apoderarse de lo más que pueda en este mundo.',
    251: 'He tenido momentos en que mi mente se ha quedado en blanco; mis actividades se interrumpieron y yo no sabía lo que pasaba a mi alrededor.',
    252: 'A nadie le importa mucho lo que le suceda a uno.',
    253: 'Puedo ser amable con personas que hacen cosas que considero incorrectas.',
    254: 'Me gusta estar en un grupo en que los unos y los otros se hagan bromas.',
    255: 'A veces en las elecciones voto por personas acerca de quienes sé muy poco.',
    256: 'La única parte interesante del periódico es la página cómica.',
    257: 'Por lo general espero tener éxito en las cosas que hago.',
    258: 'Creo que hay un Dios.',
    259: 'Me resulta difícil empezar a hacer cualquier cosa.',
    260: 'En el colegio fui lento para aprender.',
    261: 'Si fuera artista me gustaría pintar flores.',
    262: 'No me molesta el no ser mejor parecido(a) de lo que soy.',
    263: 'Transpiro con facilidad, aún en días fríos.',
    264: 'Tengo plena confianza en mí mismo.',
    265: 'Es más seguro no confiar en nadie.',
    266: 'Una o más veces por semana me pongo muy nervioso o alterado.',
    267: 'Cuando estoy en un grupo de gente, me cuesta pensar en las cosas apropiadas de que hablar.',
    268: 'Cuando me siento abatido, algo emocionante me saca casi siempre de este estado.',
    269: 'Con facilidad puedo infundirle miedo a otros, y a veces lo hago sólo por diversión.',
    270: 'Cuando salgo de casa no me preocupo si las puertas y ventanas están bien cerradas.',
    271: 'No culpo a la persona que se aprovecha de alguien que se expone a que le ocurra tal cosa.',
    272: 'A veces estoy lleno de energía.',
    273: 'Tengo adormecidas una o varias partes de la piel.',
    274: 'Mi vista es tan buena ahora como lo ha sido por años.',
    275: 'Alguien controla mi mente.',
    276: 'Me gustan los niños.',
    277: 'A veces me ha divertido tanto la astucia de un pillo (timador) que he deseado que se salga con la suya.',
    278: 'Con frecuencia he sentido que desconocidos me miraban con ojos críticos.',
    279: 'Todos los días tomo una cantidad extraordinaria de agua.',
    280: 'La mayoría de la gente se hace de amigos porque les pueden ser útiles.',
    281: 'Casi nunca noto que me zumban o chillan los oídos.',
    282: 'De vez en cuando siento odio hacia miembros de mi familia a los que normalmente quiero.',
    283: 'Si fuera reportero me gustaría mucho escribir noticias deportivas.',
    284: 'Estoy seguro de que la gente habla de mí.',
    285: 'De vez en cuando me rió de chistes sucios.',
    286: 'Nunca estoy más contento que cuando me encuentro solo.',
    287: 'Tengo pocos temores en comparación con mis amigos.',
    288: 'Sufro de ataques de náuseas y vómitos.',
    289: 'Siempre me da rabia con la ley cuando se pone en libertad a un criminal gracias a los alegatos de un abogado astuto.',
    290: 'Trabajo bajo una tensión muy grande.',
    291: 'Una o más veces en mi vida he sentido que alguien me hace hacer cosas hipnotizándome.',
    292: 'Por lo general, no le hablo a la gente hasta que ellos no me hablen a mí.',
    293: 'Alguien ha tratado de influir en mi mente.',
    294: 'Nunca he tenido problemas con la ley.',
    295: 'El cuento “Alicia en el país de las maravillas” me gustó.',
    296: 'Tengo épocas en las que me siento muy alegre sin que exista una razón especial.',
    297: 'Me gustaría que no me perturbaran pensamientos sexuales.',
    298: 'Si varias personas se hallan en apuros, lo mejor que pueden hacer es ponerse de acuerdo sobre lo que van a decir y mantenerse firmes en ello.',
    299: 'Creo que siento más intensamente que la mayoría de la gente.',
    300: 'Nunca en mi vida me ha gustado jugar con muñecas.',
    301: 'Muchas veces la vida es una carga pesada para mí.',
    302: 'Nunca me he visto en dificultades a causa de mi conducta sexual.',
    303: 'Soy tan sensible acerca de algunos asuntos que ni siquiera puedo hablar de ellos.',
    304: 'En el colegio me era muy difícil hablar frente a la clase.',
    305: 'Aún cuando esté acompañado, me siento solo la mayor parte del tiempo.',
    306: 'Recibo todas las muestras de afecto que debo recibir.',
    307: 'Rehuso participar en algunos juegos porque no los juego bien.',
    308: 'A veces he sentido un intenso deseo de abandonar mi hogar.',
    309: 'Creo que hago amistades tan fácilmente como los demás.',
    310: 'Mi vida sexual es satisfactoria.',
    311: 'Por un tiempo, cuando era más chico, me robé cosas sin importancia.',
    312: 'No me gusta tener gente alrededor.',
    313: 'El hombre que provoca la tentación dejando propiedad de valor sin protección, es tan culpable del robo como el ladrón mismo.',
    314: 'De vez en cuando pienso en cosas demasiado malas como para hablar de ellas.',
    315: 'Estoy seguro que la vida me trata mal.',
    316: 'Creo que casi todo el mundo diría una mentira para evitarse un problema.',
    317: 'Soy más sensible que la mayoría de la gente.',
    318: 'Mi vida diaria está llena de cosas que me mantienen interesado.',
    319: 'A la mayor parte de la gente, en su fuero interno, le disgusta esforzarse para ayudar a los demás.',
    320: 'Muchos de mis sueños son sobre sexo.',
    321: 'Me siento avergonzado(a) con facilidad.',
    322: 'El dinero y los negocios me preocupan.',
    323: 'He tenido experiencias muy peculiares y extrañas.',
    324: 'Nunca he estado enamorado de nadie.',
    325: 'Ciertas cosas que han hecho algunos de mis familiares me han molestado.',
    326: 'A veces me dan ataques de risa o de llanto que no puedo controlar.',
    327: 'Mi madre o mi padre frecuentemente me obligan a obedecer, aun cuando yo creía que no tenían razón.',
    328: 'Encuentro difícil concentrarme en una tarea o trabajo.',
    329: 'Casi nunca sueño.',
    330: 'Nunca he quedado paralizado ni he sufrido de ninguna debilidad muscular extraña.',
    331: 'Si la gente no la hubiera agarrado conmigo, yo habría tenido mucho más éxito.',
    332: 'Algunas veces quedo afónico o me cambia la voz, aunque no esté resfriado.',
    333: 'Nadie parece comprenderme.',
    334: 'A veces percibo olores raros.',
    335: 'No puedo concentrarme en una sola cosa.',
    336: 'Me impaciento con la gente con facilidad.',
    337: 'Siento angustia por algo o por alguien la mayor parte del tiempo.',
    338: 'Ciertamente, en cuanto a preocupaciones, me ha llovido sobre mojado.',
    339: 'La mayor parte del tiempo desearía estar muerto.',
    340: 'A veces me siento tan alterado(a) que me cuesta dominarme.',
    341: 'A veces oigo tan bien que me molesta.',
    342: 'Se me olvida inmediatamente lo que la gente me dice.',
    343: 'Generalmente tengo que detenerme a pensar antes de hacer algo, aunque sean asuntos sin importancia.',
    344: 'Con frecuencia cruzo la calle para evitar encontrarme con alguien que veo venir.',
    345: 'Muchas veces siento como si las cosas no fueran reales.',
    346: 'Tengo la costumbre de contar cosas sin importancia como ampolletas en anuncios luminosos, etc.',
    347: 'No tengo enemigos que realmente quieran hacerme daño.',
    348: 'Generalmente las personas que son un poco más amistosas de lo que yo esperaba me ponen en guardia.',
    349: 'Tengo pensamientos extraños y peculiares.',
    350: 'Oigo cosas extrañas cuando estoy solo.',
    351: 'Me angustio y me altero cuando tengo que salir de la casa para hacer un corto viaje.',
    352: 'He tenido miedo a cosas o personas que sabía que no me podían hacer daño.',
    353: 'No me gusta entrar solo a un recinto donde ya hay gente reunida hablando.',
    354: 'Tengo miedo de usar un cuchillo o cualquier otra cosa muy afilada o puntiaguda.',
    355: 'Algunas veces me gusta herir a las personas a quienes amo.',
    356: 'Tengo más dificultades para concentrarme que la que parecen tener los demás.',
    357: 'Varias veces he dejado de hacer algo porque me he creído poco capaz (incapaz).',
    358: 'Malas palabras, a menudo palabras horribles, vienen a mi mente, y se me hace imposible librarme de ellas.',
    359: 'Algunas veces se me viene a la cabeza un pensamiento sin importancia que me da vueltas días y días.',
    360: 'Casi todos los días sucede algo que me asusta.',
    361: 'Tiendo a tomar las cosas muy en serio.',
    362: 'Soy más sensible que la mayoría de la gente.',
    363: 'A veces he sentido placer cuando un ser querido me ha lastimado.',
    364: 'La gente dice cosas insultantes y vulgares acerca de mí.',
    365: 'Me siento incómodo en lugares cerrados.',
    366: 'Aún cuando este acompañado, me siento solo la mayor parte del tiempo.',
    367: 'No le temo al fuego.',
    368: 'A veces me he alejado de otra persona porque he temido hacer o decir algo de lo que pudiese arrepentirme después.',
    369: 'La religión no me causa problemas.',
    370: 'Me carga tener que apurarme cuando trabajo.',
    371: 'No soy una persona demasiado cohibida o tímida.',
    372: 'Tiendo a interesarme en varios hobbies en vez de concentrarme por largo tiempo en uno de ellos.',
    373: 'Estoy seguro de que sólo existe una religión verdadera.',
    374: 'Durante ciertos períodos, mi mente parece trabajar más despacio que de costumbre.',
    375: 'Cuando me siento muy feliz y activo, alguien que esté triste o deprimido, me lo echa a perder todo.',
    376: 'Los policías son generalmente honrados.',
    377: 'En las reuniones sociales o fiestas, es más probable que me siente solo o con una sola persona, en vez de unirme al grupo.',
    378: 'No me gusta ver fumar a las mujeres.',
    379: 'Muy rara vez me siento melancólico.',
    380: 'Cuando alguien dice cosas tontas o que denotan ignorancia acerca de algo que yo sé, trato de corregirlo.',
    381: 'Con frecuencia me han dicho que soy arrebatado (impulsivo).',
    382: 'Quisiera dejar de preocuparme por cosas que he dicho, y que quizá hayan herido los sentimientos de otros.',
    383: 'La gente me desilusiona con frecuencia.',
    384: 'Me siento incapaz de contarle a alguien todas mis cosas.',
    385: 'Me dan miedo los relámpagos.',
    386: 'Me gusta tener a los demás intrigados acerca de lo que voy a hacer.',
    387: 'Los únicos milagros que conozco son simplemente tretas que unas personas le hacen a otras.',
    388: 'Me da miedo estar solo(a) en la oscuridad.',
    389: 'Con frecuencia mis planes han parecido estar tan llenos de dificultades, que he tenido que abandonarlos.',
    390: 'Muchas veces me he sentido muy mal al ser mal interpretado cuando trataba de evitar que alguien cometiese un error.',
    391: 'Me gusta mucho ir a bailes.',
    392: 'Le tengo terror a las tormentas de viento.',
    393: 'Los caballos que no tiran debieran ser golpeados o pateados.',
    394: 'Frecuentemente pido consejo a la gente.',
    395: 'El futuro es demasiado incierto para que una persona haga planes serios.',
    396: 'Con frecuencia, aún cuando todo me sale bien, siento que nada me importa.',
    397: 'Algunas veces he sentido que las dificultades se acumularon de tal modo que no he podido superarlo.',
    398: 'A menudo pienso: “quisiera volver a ser niño”.',
    399: 'No me enojo fácilmente.',
    400: 'Si me dieran la oportunidad, podría hacer algunas cosas que serían de gran beneficio para la humanidad.',
    401: 'No le temo al agua.',
    402: 'Frecuentemente tengo que consultar con la almohada antes de tomar decisiones.',
    403: 'Es una gran cosa vivir en esta época que ocurren tantas cosas.',
    404: 'A menudo la gente ha interpretado mal mis intenciones cuando trataba de corregirlo o ayudarla.',
    405: 'No tengo dificultad al tragar.',
    406: 'A menudo he conocido personas a quienes se suponía expertos y que no eran mejores que yo.',
    407: 'Por lo general soy tranquilo y no me altero fácilmente.',
    408: 'Puedo ocultar lo que siento en algunas cosas, hasta el punto en que la gente puede herirme sin que se den cuenta de ello.',
    409: 'A veces me he agotado por emprender demasiadas cosas.',
    410: 'Me encantaría ganarle a un pillo (timador) con sus propias armas.',
    411: 'Me siento un fracasado cuando oigo hablar del éxito de alguien a quien conozco bien.',
    412: 'No me asusta consultar al médico por una enfermedad o lesión.',
    413: 'Merezco un severo castigo por mis pecados.',
    414: 'Tiendo a preocuparme tanto por los desengaños, que luego no puedo dejar de pensar en ellos.',
    415: 'Si me dieran la oportunidad sería un buen líder.',
    416: 'Me molesta que alguien me observe cuando trabajo, aunque sepa que puedo hacerlo bien.',
    417: 'A menudo me siento tan molesto cuando alguien trata de adelantárseme en una cola que le llamo la atención.',
    418: 'A veces pienso que no sirvo para nada.',
    419: 'Cuando estaba en el colegio, ”hice la cimarra” muy a menudo.',
    420: 'He tenido experiencias religiosas muy poco corrientes.',
    421: 'Tengo uno o varios miembros de mi familia que son muy nerviosos.',
    422: 'Me he sentido avergonzado por la clase de trabajo que alguien de mi familia ha hecho.',
    423: 'Me gusta o me ha gustado muchísimo pescar.',
    424: 'Siento hambre casi todo el tiempo.',
    425: 'Sueño frecuentemente.',
    426: 'A veces he tenido que ser rudo con personas groseras o inoportunas.',
    427: 'Me avergüenzan los chistes groseros.',
    428: 'Me gusta leer los editoriales de los periódicos.',
    429: 'Me gusta asistir a conferencias sobre temas serios.',
    430: 'Me atraen las personas del sexo opuesto.',
    431: 'Me preocupo bastante por posibles desgracias.',
    432: 'Tengo opiniones bien firmes.',
    433: 'Acostumbro a tener amigos imaginarios.',
    434: 'Me gustaría competir en carreras automovilísticas.',
    435: 'Generalmente preferiría trabajar con mujeres.',
    436: 'Generalmente la gente exige más respecto a sus propios derechos, que lo que está dispuesto a respetar los de los demás.',
    437: 'No es malo tratar de evitar el cumplimiento de la ley, siempre que ésta no se viole.',
    438: 'Hay ciertas personas que me disgustan tanto, que me alegro interiormente cuando están pagando las consecuencias por algo que han hecho.',
    439: 'Me pone nervioso tener que esperar.',
    440: 'Trato de recordar anécdotas interesantes para contárselas a otras personas.',
    441: 'Me gustan las mujeres altas.',
    442: 'He tenido períodos durante los cuales he perdido el sueño a causa de las preocupaciones.',
    443: 'Tiendo a dejar de hacer algo que deseo hacer cuando otros piensan que no lo estoy haciendo bien.',
    444: 'No trato de corregir a la gente que expresa creencias ignorantes.',
    445: 'Cuando era joven o en mi niñez, me apasionaba lo emocionante.',
    446: 'Me gusta apostar cuando se trata de poco dinero.',
    447: 'Con frecuencia me esfuerzo para triunfar sobre alguien que me ha llevado la contraria.',
    448: 'Me molesta que la gente en las tiendas, buses, etc., me esté mirando.',
    449: 'Me gustan las reuniones sociales sólo por estar con la gente.',
    450: 'Me gusta la excitación de las multitudes.',
    451: 'Mis preocupaciones parece que desaparecen cuando estoy con un grupo de amigos entusiastas.',
    452: 'Me gusta burlarme de la gente.',
    453: 'Cuando era niño nunca me interesó pertenecer a un grupo o pandilla.',
    454: 'Podría ser feliz viviendo completamente solo en una cabaña, en el bosque o en las montañas.',
    455: 'Frecuentemente no me entero de los chismes y habladurías del grupo al que pertenezco.',
    456: 'Una persona no debiera ser castigada por violar una ley que considera poco injusta.',
    457: 'Creo que no debería probar nunca bebidas alcohólicas.',
    458: 'El hombre que más tenía que ver conmigo cuando era niño(a) (como mi padre, padrastro, etc.) fue muy estricto.',
    459: 'Tengo uno o varios malos hábitos tan arraigados que es inútil luchar contra ellos.',
    460: 'He bebido alcohol moderadamente (o nunca lo he tomado).',
    461: 'Me es difícil dejar a un lado una tarea que he emprendido, aún cuando sea por poco tiempo.',
    462: 'No he tenido dificultad para empezar a orinar o retener mi orina.',
    463: 'Me gustaba jugar al “luche”.',
    464: 'Nunca he tenido una visión.',
    465: 'Varias veces me he arrepentido de cómo he encauzado mi trabajo.',
    466: 'Excepto por orden del médico, nunca tomo drogas o pastillas para dormir.',
    467: 'Con frecuencia memorizo números sin importancia (tales como los de las patentes de automóviles, etc.).',
    468: 'Frecuentemente me siento apenado por ser tan mal genio y gruñón.',
    469: 'A menudo me he dado cuenta que había gente envidiosa de mis buenas ideas sólo por que no se les había ocurrido antes.',
    470: 'Me disgustan las cosas sexuales.',
    471: 'En el colegio mis notas de conducta fueron generalmente malas.',
    472: 'Me fascina el fuego.',
    473: 'Siempre que me es posible evito encontrarme frente a una multitud.',
    474: 'No tengo que orinar con más frecuencia que los demás.',
    475: 'Cuando estoy acorralado digo sólo aquella parte de la verdad que no me perjudica.',
    476: 'Soy un enviado especial de Dios.',
    477: 'Si me hallara en dificultades junto con varios amigos que fueran tan culpables como yo, preferiría echarme toda la culpa antes de descubrirlos.',
    478: 'Nunca me he puesto particularmente nervioso a causa de las dificultades en que se haya visto envuelto algún miembro de mi familia.',
    479: 'No me molesta el conocer gente nueva.',
    480: 'Con frecuencia le tengo miedo a la oscuridad.',
    481: 'Recuerdo haberme sentido enfermo para zafarme de algo.',
    482: 'En los trenes, buses, etc. con frecuencia converso con desconocidos.',
    483: 'Cristo realizó milagros tales como cambiar el agua en vino.',
    484: 'Tengo uno o más defectos que son tan grandes que es mejor aceptarlos y tratar de controlarlos, antes de tratar de liberarme de ellos.',
    485: 'Cuando un hombre está con una mujer generalmente está pensando en cosas relativas al sexo de ella.',
    486: 'Nunca he notado sangre en mi orina.',
    487: 'Me rindo fácilmente cuando las cosas me salen mal.',
    488: 'Rezo varias veces a la semana.',
    489: 'Me compadezco de las personas que se aferran a sus penas y problemas.',
    490: 'Leo la Biblia varias veces a la semana.',
    491: 'No tolero a la gente que cree que sólo hay una religión verdadera.',
    492: 'Me produce terror la idea de un terremoto.',
    493: 'Prefiero el trabajo que requiere concentración a un trabajo que me permite ser descuidado.',
    494: 'Temo encontrarme en un closet o lugar pequeño y cerrado.',
    495: 'Generalmente “le hablo claro a la gente a quien estoy tratando de mejorar o corregir”.',
    496: 'Nunca he visto las cosas dobles (es decir, nunca un objeto me ha parecido doble sin que me sea posible verlo como uno).',
    497: 'Me gustan las historias de aventuras.',
    498: 'Siempre es bueno ser franco.',
    499: 'Tengo que admitir que a veces me he preocupado sin motivo alguno por cosas que no valían la pena.',
    500: 'Fácilmente me vuelvo partidario(a) absoluto de una buena idea.',
    501: 'Generalmente resuelvo las cosas solo, sin buscar a alguien que me enseñe.',
    502: 'Me gusta hacerle saber a la gente lo que pienso acerca de las cosas.',
    503: 'Es raro que yo apruebe o desapruebe con energía las acciones de otros.',
    504: 'No trato de esconder la mala opinión o lástima que me inspira una persona, a fin de que ésta no sepa mi manera de sentir.',
    505: 'He tenido períodos en que me sentía tan lleno de energía que me parecía que no necesitaba dormir a ninguna hora.',
    506: 'Soy una persona muy tensa.',
    507: 'Frecuentemente he trabajado para personas que parece que arreglan las cosas de tal modo, que ellas son las que reciben el reconocimiento de una buena labor, pero que sin embargo atribuyen los errores a los otros que están bajo ellos.',
    508: 'Creo que mi olfato es tan bueno como el de los demás.',
    509: 'A veces me es difícil defender mis derechos por ser muy reservado.',
    510: 'La mugre me espanta o me da asco.',
    511: 'Vivo una vida de ensueños acerca de la cual no digo nada a nadie.',
    512: 'No me gusta bañarme.',
    513: 'Creo que Cervantes fue más grande que Napoleón.',
    514: 'Me gustan las mujeres “ahombradas”.',
    515: 'En mi hogar siempre hemos tenido cubiertas nuestras necesidades básicas (tales como alimentación, vestuario, etc.).',
    516: 'Algunos de mis familiares tienen mal genio.',
    517: 'No puedo hacer nada bien.',
    518: 'A menudo me he sentido culpable porque he sentido mayor pesar del que realmente sentía.',
    519: 'Algo anda mal con mis órganos sexuales.',
    520: 'Como norma, defiendo con firmeza mis propias opiniones.',
    521: 'No me turbaría ante un grupo de personas, si tuviera que iniciar una discusión o dar una opinión de algo que conozco bien.',
    522: 'No le temo a las arañas.',
    523: 'Casi nunca me ruborizo.',
    524: 'No temo contagiarme o coger gérmenes de las perillas de las puertas.',
    525: 'Ciertos animales me ponen nervioso.',
    526: 'El porvenir me parece sin esperanzas.',
    527: 'Los miembros de mi familia y mis parientes más cercanos se llevan bastante bien.',
    528: 'No me ruborizo con mayor frecuencia que los demás.',
    529: 'Me gustaría usar ropa fina (de buena calidad, cara).',
    530: 'A menudo siento miedo de ruborizarme.',
    531: 'La gente puede hacerme cambiar de opinión muy fácilmente, aún en cosas en las que creía estar ya decidido.',
    532: 'Puedo soportar tanto dolor como los demás.',
    533: 'No padezco de mucha flatulencia.',
    534: 'Varias veces he sido el último en darme por vencido al tratar de hacer algo.',
    535: 'Siento la boca seca casi todo el tiempo.',
    536: 'Me molesta que la gente me apure.',
    537: 'Me gustaría cazar leones en África.',
    538: 'Creo que me gustaría el trabajo de modista (o modista).',
    539: 'No le tengo miedo a los ratones.',
    540: 'Nunca he sufrido de parálisis facial.',
    541: 'Mi piel parece ser extraordinariamente sensible al tacto.',
    542: 'Nunca he tenido deposiciones (excretas) negras, parecidas a la brea.',
    543: 'Varias veces por semana siento como si algo terrible fuera a suceder.',
    544: 'La mayor parte del tiempo me siento cansado (a).',
    545: 'Algunas veces sueño lo mismo una y otra vez.',
    546: 'Me gusta leer sobre historia.',
    547: 'Me gustan las fiestas y las reuniones sociales.',
    548: 'Nunca asisto a un espectáculo frívolo si es que puedo evitarlo.',
    549: 'Me acobardo al enfrentar una crisis o dificultad.',
    550: 'Me gusta reparar las cerraduras de las puertas.',
    551: 'A veces estoy seguro(a) que los demás saben lo que estoy pensando.',
    552: 'Me gusta leer sobre ciencia.',
    553: 'Tengo miedo de estar solo en un sitio amplio al descubierto.',
    554: 'Si fuera artista me gustaría dibujar niños.',
    555: 'Algunas veces me siento a punto de derrumbarme.',
    556: 'Soy muy cuidadoso en mi manera de vestir.',
    557: 'Me gustaría ser secretario(a) privado(a).',
    558: 'Un gran número de personas son culpables de mala conducta sexual.',
    559: 'Con frecuencia he sentido miedo en la noche.',
    560: 'Me molesta mucho que se me olvide donde pongo las cosas.',
    561: 'Me gusta mucho montar a caballo.',
    562: 'La persona hacia quien sentía mayor afecto y admiración cuando niño era una mujer (madre, hermana, tía u otra mujer).',
    563: 'Me gustan más las historias de aventura que las de amor.',
    564: 'Puedo dejar de hacer algo que deseo (hacer) cuando otros creen que no vale la pena hacerlo.',
    565: 'Siento deseos de tirarme cuando estoy en un sitio alto.',
    566: 'Me gustan las escenas de amor en las películas.'
}


# 用于存放测验的原始结果
# Used to store the original results of the questionnaire
Ans = {

}


def start():
    """
    Sección de la guía de preguntas con introducción y directrices
    Guide section, including abstract and instruction

    :return: None
    """
    print(Ins1)
    time.sleep(5)

    while 1:
        print('El cuestionario empezará a continuación, ¿empezamos? ')
        print('1. Sí    0. No')
        go = input('> ')
        if go == '1':
            print('-' * 65)
            print('El cuestionario comenzará oficialmente en 30 segundos, por favor lea y entienda lo siguiente cuidadosamente')
            time.sleep(3)
            print(Ins2)
            time.sleep(27)
            break
        elif go == '0':
            print('Gracias por utilizar este programa, ¡adiós!')
            time.sleep(3)
            exit(0)
        else:
            print('Si se ha equivocado, vuelva a introducir los datos tal y como exige el cuestionario.')
            continue


def answer():
    """for debug
    Preguntas y respuestas aleatorias
    random answer

    :return: '0' or '1'
    :rtype: str
    """
    ans = randint(0, 1)
    return str(ans)


def test():
    """
    Cuestionario
    test section

    :return: None
    """
    global Sex
    global Age

    print('El cuestionario ha comenzado oficialmente. ')
    print('-' * 65)
    time.sleep(3)

    while 1:
        print('x1. Mi género es')
        print('1. 男    0.女')
        Sex = input('> ')
        if Sex == '1' or Sex == '0':
            break
        else:
            print('输入错误请按照测验要求重新输入！')
            continue

    while 1:
        print('-' * 65)
        print('x2. 请输入你的年龄')
        Age = input('> ')

        if str.isdigit(Age):
            if 13 <= int(Age) <= 70:
                break
            else:
                print('本测验不适用于该年龄范围，感谢使用！')
                time.sleep(3)
                exit(0)
        else:
            print('输入错误请按照测验要求重新输入！')
            continue

    for i in range(len(Que)+1):
        if i == 73:
            if Sex == '1':
                temp_que = str(i+1) + '. ' + (Que[i+1][Que[i+1].find('m')+1: Que[i+1].find('f')])
            else:
                temp_que = str(i+1) + '. ' + (Que[i+1][Que[i+1].find('f') + 1:])
        elif i == len(Que):
            temp_que = str(len(Que)+1) + '.' + '我保证是在专业人士指导下认真诚实地完成本次测验'
        else:
            temp_que = str(i+1) + '. ' + Que[i+1]

        while 1:
            print('-' * 65)
            print(temp_que)
            print('1. 是    0. 否')
            temp_ans = input('> ')
            # temp_ans = answer()  # for debug
            # print('> ' + str(temp_ans))  # for debug
            if temp_ans == '1' or temp_ans == '0':
                Ans[i+1] = temp_ans
                # print(Ans)  # for debug
                break
            elif temp_ans == 'bomb':    # for debug
                exit(0)
            else:
                print('输入错误请按照测验要求重新输入！')
                continue

    print('-' * 65)
    print('测验结束，感谢您的配合！')
    print('-' * 65)


def is_diff(a, b):
    """
    为异计分
    Add point if different

    :param a: the first para
    :param b: the second para
    :return: 0 or 1

    :type a: str
    :type b: str
    :rtype: int
    """
    if a != b:
        return 1
    else:
        return 0


def is_true(t):
    """
    正向计分
    Add point if True

    :param t: the para under test
    :return: 0 or 1

    :type t: str
    :rtype: int
    """
    if t == '1':
        return 1
    else:
        return 0


def is_false(t):
    """
    反向计分
    Add point if False

    :param t: the para under test
    :return: 0 or 1

    :type t: str
    :rtype: 0 or 1
    """
    if t == '0':
        return 1
    else:
        return 0


def norm_select(sex):
    """
    选择常模表（中国1982版）
    the norm select (based on Chinese 1982's)

    :param sex: the subjects' sex
    :return: Norm_M, Norm_SD

    :type sex: str
    :rtype: None
    """
    global Norm_M
    global Norm_SD

    # 男性常模
    # male norm
    if sex == '1':
        Norm_M = {
            'L': 5.70,
            'F': 13.68,
            'K': 13.00,
            'Hs': 8.78,
            'D': 26.16,
            'Hy': 22.07,
            'Pd': 18.98,
            'Mf': 27.56,
            'Pa': 12.84,
            'Pt': 17.86,
            'Sc': 23.01,
            'Ma': 18.48,
            'Si': 34.51,
            'Hs+0.5K': 15.42,
            'Pd+0.4K': 24.38,
            'Pt+1.0K': 31.14,
            'Sc+1.0K': 36.47,
            'Ma+0.2K': 21.22,
            'Mas': 18.86,
            'Dy': 26.09,
            'Do': 15.39,
            'Re': 20.54,
            'Cn': 25.26
        }
        Norm_SD = {
            'L': 2.52,
            'F': 6.86,
            'K': 4.66,
            'Hs': 4.75,
            'D': 4.97,
            'Hy': 5.36,
            'Pd': 4.36,
            'Mf': 4.04,
            'Pa': 3.92,
            'Pt': 7.93,
            'Sc': 10.15,
            'Ma': 5.26,
            'Si': 6.88,
            'Hs+0.5K': 4.79,
            'Pd+0.4K': 4.27,
            'Pt+1.0K': 5.71,
            'Sc+1.0K': 8.24,
            'Ma+0.2K': 4.88,
            'Mas': 7.45,
            'Dy': 8.05,
            'Do': 3.12,
            'Re': 4.13,
            'Cn': 3.76
        }
    # 女性常模
    # female norm
    else:
        Norm_M = {
            'L': 5.64,
            'F': 11.69,
            'K': 12.25,
            'Hs': 9.83,
            'D': 28.40,
            'Hy': 22.82,
            'Pd': 18.29,
            'Mf': 31.83,
            'Pa': 12.62,
            'Pt': 18.77,
            'Sc': 22.50,
            'Ma': 16.64,
            'Si': 37.27,
            'Hs+0.5K': 16.35,
            'Pd+0.4K': 23.33,
            'Pt+1.0K': 31.17,
            'Sc+1.0K': 34.89,
            'Ma+0.2K': 19.18,
            'Mas': 20.43,
            'Dy': 29.12,
            'Do': 15.10,
            'Re': 21.78,
            'Cn': 24.86
        }
        Norm_SD = {
            'L': 2.48,
            'F': 5.02,
            'K': 4.26,
            'Hs': 4.98,
            'D': 5.04,
            'Hy': 5.54,
            'Pd': 4.45,
            'Mf': 3.86,
            'Pa': 3.93,
            'Pt': 7.82,
            'Sc': 9.57,
            'Ma': 5.16,
            'Si': 6.71,
            'Hs+0.5K': 4.95,
            'Pd+0.4K': 4.44,
            'Pt+1.0K': 5.86,
            'Sc+1.0K': 7.63,
            'Ma+0.2K': 4.89,
            'Mas': 7.35,
            'Dy': 7.61,
            'Do': 2.76,
            'Re': 3.13,
            'Cn': 3.70
        }



def trans_t(score, m, sd):
    """
    标准T分计算公式
    Standard T point conversion formula

    :param score: original score
    :param m: normative mean value
    :param sd: normative standard deviation
    :return: standard T score

    :type score: int
    :type m: float
    :type sd: float
    :rtype: int
    """
    t = round(50 + 10*(score - m)/sd)
    return t


def scale_q(ori_score=0, pro_score=0):
    """
    效度量表-疑问分数 Q
    the score of Q (? or question) scale,

    由于不允许被试者存在空题，故仅记录16项重复问题的矛盾数量
    because subjects were not allowed to have blank questions,
    just record the number of contradictions of 16 repeated questions

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    # 原始分 original score
    temp1 = [8, 13, 15, 16, 20, 21, 22, 23, 24, 32, 33, 35, 37, 38, 305, 317]
    temp2 = [318, 290, 314, 315, 310, 308, 326, 288, 333, 328, 323, 331, 302, 311, 366, 362]

    for i in range(len(temp1)):
        ori_score += is_diff(Ans[temp1[i]], Ans[temp2[i]])

    temp = ori_score
    pro_score += temp

    return ori_score, pro_score


def scale_l(ori_score=0, pro_score=0):
    """
    效度量表-说谎分数 L
    the score of L (lie) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    # 原始分 original score
    temp = [15, 30, 45, 60, 75, 90, 105, 120, 135, 150, 165, 195, 225, 255, 285]

    for i in temp:
        ori_score += is_false(Ans[i])

    pro_score += trans_t(ori_score, Norm_M['L'], Norm_SD['L'])

    return ori_score, pro_score


def scale_f(ori_score=0, pro_score=0):
    """
    效度量表-诈病分数 F
    the score of F (infrequency or fake bad) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    # 原始分 original score
    temp_t = [14, 27, 31, 34, 35, 40, 42, 48, 49, 50, 53, 56, 66, 85, 121, 123, 139, 146, 151, 156, 168, 184, 197, 200,
              202, 205, 206, 209, 210, 211, 215, 218, 227, 245, 246, 247, 252, 256, 269, 275, 286, 288, 291, 293]
    temp_f = [17, 20, 54, 65, 75, 83, 112, 113, 115, 164, 169, 177, 185, 196, 199, 220, 257, 258, 272, 276]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['F'], Norm_SD['F'])

    return ori_score, pro_score


def scale_k(ori_score=0, pro_score=0):
    """
    效度量表-校正分数 K
    the score of K (defensiveness) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [96]
    temp_f = [30, 39, 71, 89, 124, 129, 134, 138, 142, 148, 160, 170, 171, 180, 183, 217, 234, 267, 272, 296, 316, 322,
              368, 370, 372, 373, 375, 386, 394]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['K'], Norm_SD['K'])

    return ori_score, pro_score


def scale_hs(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    临床量表-1 疑病 Hs
    the score of Hs (hypochondriasis) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 0.5K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [23, 29, 43, 62, 72, 108, 114, 125, 161, 189, 273]
    temp_f = [2, 3, 7, 9, 18, 51, 55, 63, 68, 103, 130, 153, 155, 163, 175, 188, 190, 192, 230, 243, 274, 281]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Hs'], Norm_SD['Hs'])
    pro_score_add_k += trans_t(ori_score + round(0.5 * k), Norm_M['Hs+0.5K'], Norm_SD['Hs+0.5K'])

    return ori_score, pro_score, pro_score_add_k


def scale_d(ori_score=0, pro_score=0):
    """
    临床量表-2 抑郁 D
    the score of D (depression) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [5, 32, 41, 43, 52, 67, 86, 104, 130, 138, 142, 158, 159, 182, 189, 193, 236, 259, 288, 290]
    temp_f = [2, 8, 9, 18, 30, 36, 39, 46, 51, 57, 58, 64, 80, 88, 89, 95, 98, 107, 122, 131, 145, 152, 153, 154, 155,
              160, 178, 191, 207, 208, 233, 241, 242, 248, 263, 270, 271, 272, 285, 296]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['D'], Norm_SD['D'])

    return ori_score, pro_score


def scale_hy(ori_score=0, pro_score=0):
    """
    临床量表-3 癔病 Hy
    the score of Hy (hysteria) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [10, 23, 32, 43, 44, 47, 76, 114, 179, 186, 189, 238, 253]
    temp_f = [2, 3, 6, 7, 8, 9, 12, 26, 30, 51, 55, 71, 89, 93, 103, 107, 109, 124, 128, 129, 136, 137, 141, 147, 153,
              160, 162, 163, 170, 172, 174, 175, 180, 188, 190, 192, 201, 213, 230, 234, 243, 265, 267, 274, 279, 289,
              292]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Hy'], Norm_SD['Hy'])

    return ori_score, pro_score


def scale_pd(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    临床量表-4 精神病态 Pd
    the score of Pd (psychopathic deviate) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 0.4K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [16, 21, 24, 32, 33, 35, 38, 42, 61, 67, 84, 94, 102, 106, 110, 118, 127, 215, 216, 224, 239, 244, 245,
              284]
    temp_f = [8, 20, 37, 82, 91, 96, 107, 134, 137, 141, 155, 170, 171, 173, 180, 183, 201, 231, 235, 237, 248, 267,
              287, 289, 294, 296]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Pd'], Norm_SD['Pd'])
    pro_score_add_k += trans_t(ori_score + round(0.4 * k), Norm_M['Pd+0.4K'], Norm_SD['Pd+0.4K'])

    return ori_score, pro_score, pro_score_add_k


def scale_mf(ori_score=0, pro_score=0):
    """
    临床量表-5 男子气/女子气 Mf
    the score of Mf (masculinity-femininity) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    # 男性女性化 Mf-m
    if Sex == '1':
        temp_t = [4, 25, 69, 70, 74, 77, 78, 87, 92, 126, 132, 134, 140, 149, 179, 187, 203, 204, 217, 226, 231, 239,
                  261, 278, 282, 295, 297, 299]
        temp_f = [1, 19, 26, 28, 79, 80, 81, 89, 99, 112, 115, 116, 117, 120, 133, 144, 176, 198, 213, 214, 219, 221,
                  223, 229, 249, 254, 260, 262, 264, 280, 283, 300]
    # 女性男性化 Mf-f
    else:
        temp_t = [4, 25, 70, 74, 77, 78, 87, 92, 126, 132, 133, 134, 140, 149, 187, 203, 204, 217, 226, 239, 261, 278,
                  282, 295, 299]
        temp_f = [1, 19, 26, 28, 69, 79, 80, 81, 89, 99, 112, 115, 116, 117, 120, 144, 176, 179, 198, 213, 214, 219,
                  221, 223, 229, 231, 249, 254, 260, 262, 264, 280, 283, 297, 300]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Mf'], Norm_SD['Mf'])

    return ori_score, pro_score


def scale_pa(ori_score=0, pro_score=0):
    """
    临床量表-6 妄想狂 Pa
    the score of Pa (paranoia) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [16, 24, 27, 35, 110, 121, 123, 127, 151, 157, 158, 202, 275, 284, 291, 293, 299, 305, 314, 317, 326, 338,
              341, 364, 365]
    temp_f = [93, 107, 109, 111, 117, 124, 268, 281, 294, 313, 316, 319, 327, 347, 348]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Pa'], Norm_SD['Pa'])

    return ori_score, pro_score


def scale_pt(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    临床量表-7 精神衰弱 Pt
    the score of Pt (psychasthenia) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 1.0K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [10, 15, 22, 32, 41, 67, 76, 86, 94, 102, 106, 142, 159, 182, 189, 217, 238, 266, 301, 304, 321, 336, 337,
              340, 342, 343, 344, 346, 349, 351, 352, 356, 357, 358, 359, 360, 361, 362, 366]
    temp_f = [3, 8, 36, 122, 152, 164, 178, 329, 353]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Pt'], Norm_SD['Pt'])
    pro_score_add_k += trans_t(ori_score + k, Norm_M['Pt+1.0K'], Norm_SD['Pt+1.0K'])

    return ori_score, pro_score, pro_score_add_k


def scale_sc(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    临床量表-8 精神分裂症 Sc
    the score of Sc (schizophrenia) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 1.0K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [15, 22, 40, 41, 47, 52, 76, 97, 104, 121, 156, 157, 159, 168, 179, 182, 194, 202, 210, 212, 238, 241, 251,
              259, 266, 273, 282, 291, 297, 301, 303, 307, 308, 311, 312, 315, 320, 323, 324, 325, 328, 331, 332, 333,
              334, 335, 339, 341, 345, 349, 350, 352, 354, 355, 356, 360, 363, 364, 366]
    temp_f = [17, 65, 103, 119, 177, 178, 187, 192, 196, 220, 276, 281, 302, 306, 309, 310, 318, 322, 330]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Sc'], Norm_SD['Sc'])
    pro_score_add_k += trans_t(ori_score + k, Norm_M['Sc+1.0K'], Norm_SD['Sc+1.0K'])

    return ori_score, pro_score, pro_score_add_k


def scale_ma(ori_score=0, pro_score=0, pro_score_add_k=0):
    """
    临床量表-9 轻躁狂 Ma
    the score of Ma (hypomania) scale

    :param ori_score: original score
    :param pro_score: processing score
    :param pro_score_add_k: processing score added 0.2K
    :return: ori_score, pro_score, pro_score_add_k

    :rtype: int, int, int
    """
    temp_t = [11, 13, 21, 22, 59, 64, 73, 97, 100, 109, 127, 134, 143, 156, 157, 167, 181, 194, 212, 222, 226, 228, 232,
              233, 238, 240, 250, 251, 263, 266, 268, 271, 277, 279, 298]
    temp_f = [101, 105, 111, 119, 120, 148, 166, 171, 180, 267, 289]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    k, ignore = scale_k()
    pro_score += trans_t(ori_score, Norm_M['Ma'], Norm_SD['Ma'])
    pro_score_add_k += trans_t(ori_score + round(0.2 * k), Norm_M['Ma+0.2K'], Norm_SD['Ma+0.2K'])

    return ori_score, pro_score, pro_score_add_k


def scale_si(ori_score=0, pro_score=0):
    """
    临床量表-0 社会内向性 Si
    the score of Si (social introversion) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [32, 67, 82, 111, 117, 124, 138, 147, 171, 172, 180, 201, 236, 267, 278, 292, 304, 316, 321, 332, 336, 342,
              357, 369, 370, 373, 376, 378, 379, 385, 389, 393, 398, 399]
    temp_f = [25, 33, 57, 91, 99, 110, 126, 143, 193, 208, 229, 231, 254, 262, 281, 296, 309, 353, 359, 367, 371, 374,
              377, 380, 381, 382, 383, 384, 387, 388, 390, 391, 392, 395, 396, 397]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Si'], Norm_SD['Si'])

    return ori_score, pro_score


def scale_mas(ori_score=0, pro_score=0):
    """
    附加量表- 外显性焦虑 Mas
    the score of Mas (Manifest anxiety) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [13, 14, 23, 31, 32, 43, 67, 86, 125, 142, 158, 186, 191, 217, 238, 241, 263, 301, 317, 321, 322, 335, 337,
              340, 352, 361, 372, 398, 418, 424, 431, 439, 442, 499, 506, 530, 555]
    temp_f = [7, 18, 107, 163, 190, 230, 242, 264, 287, 367, 407, 520, 528]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Mas'], Norm_SD['Mas'])

    return ori_score, pro_score


def scale_dy(ori_score=0, pro_score=0):
    """
    附加量表- 依赖性 Dy
    the score of Dy (Dependency) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [19, 21, 24, 41, 63, 67, 70, 82, 86, 98, 100, 138, 141, 158, 165, 180, 189, 201, 212, 236, 239, 259, 267,
              304, 305, 321, 337, 338, 343, 357, 361, 362, 370, 372, 373, 393, 398, 399, 408, 440, 443, 461, 487, 488,
              489, 509, 521, 531, 554]
    temp_f = [9, 79, 107, 163, 170, 193, 264, 411]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Dy'], Norm_SD['Dy'])

    return ori_score, pro_score


def scale_do(ori_score=0, pro_score=0):
    """
    附加量表- 支配性 Do
    the score of Do (Dominance) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [64, 229, 255, 270, 406, 432, 523]
    temp_f = [32, 61, 82, 86, 94, 186, 223, 224, 240, 249, 250, 267, 268, 304, 343, 356, 419, 483, 547, 558, 562]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Do'], Norm_SD['Do'])

    return ori_score, pro_score


def scale_re(ori_score=0, pro_score=0):
    """
    附加量表- 社会责任感 Re
    the score of Re (Social Responsibility) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [58, 111, 173, 221, 294, 412, 501, 552]
    temp_f = [6, 28, 30, 33, 56, 116, 118, 157, 175, 181, 223, 224, 260, 304, 388, 419, 434, 437, 468, 471, 472, 529,
              553, 558]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Re'], Norm_SD['Re'])

    return ori_score, pro_score


def scale_cn(ori_score=0, pro_score=0):
    """
    附加量表- 控制 Cn
    the score of Cn (Control) scale

    :param ori_score: original score
    :param pro_score: processing score
    :return: ori_score, pro_score

    :rtype: int, int
    """
    temp_t = [6, 20, 30, 56, 67, 105, 116, 134, 145, 162, 169, 181, 225, 236, 238, 285, 296, 319, 337, 376, 379, 381,
              418, 447, 460, 461, 529, 555]
    temp_f = [58, 80, 92, 96, 111, 167, 174, 220, 242, 249, 250, 291, 313, 360, 439, 444, 449, 483, 488, 489, 527, 548]

    for i in temp_t:
        ori_score += is_true(Ans[i])
    for j in temp_f:
        ori_score += is_false(Ans[j])

    pro_score += trans_t(ori_score, Norm_M['Cn'], Norm_SD['Cn'])

    return ori_score, pro_score


def calculate_score():
    """
    测验分数计算
    calculate the score

    :return: None
    """
    global ori_Point
    global pro_Point

    ori_Point = {

    }
    pro_Point = {

    }

    norm_select(Sex)

    ori_Point['Q*'], pro_Point['Q*'] = scale_q()
    ori_Point['L'], pro_Point['L'] = scale_l()
    ori_Point['F'], pro_Point['F'] = scale_f()
    ori_Point['K'], pro_Point['K'] = scale_k()
    ori_Point['Hs'], pro_Point['Hs'], pro_Point['Hs+0.5K'] = scale_hs()
    ori_Point['D'], pro_Point['D'] = scale_d()
    ori_Point['Hy'], pro_Point['Hy'] = scale_hy()
    ori_Point['Pd'], pro_Point['Pd'], pro_Point['Pd+0.4K'] = scale_pd()
    ori_Point['Mf'], pro_Point['Mf'] = scale_mf()
    ori_Point['Pa'], pro_Point['Pa'] = scale_pa()
    ori_Point['Pt'], pro_Point['Pt'], pro_Point['Pt+1.0K'] = scale_pt()
    ori_Point['Sc'], pro_Point['Sc'], pro_Point['Sc+1.0K'] = scale_sc()
    ori_Point['Ma'], pro_Point['Ma'], pro_Point['Ma+0.2K'] = scale_ma()
    ori_Point['Si'], pro_Point['Si'] = scale_si()
    ori_Point['Mas'], pro_Point['Mas'] = scale_mas()
    ori_Point['Dy'], pro_Point['Dy'] = scale_dy()
    ori_Point['Do'], pro_Point['Do'] = scale_do()
    ori_Point['Re'], pro_Point['Re'] = scale_re()
    ori_Point['Cn'], pro_Point['Cn'] = scale_cn()


def analyze_score():
    """
    分析测验分数
    analyze the score of test

    利用两点编码法以及剖析图方式呈现被测者的人格特点
    Use 2 point codes and personality profile to show the personality traits of the subjects

    :return:None
    """
    global two_point

    val_scale = ['L', 'F', 'K']
    cli_scale = ['Hs\n1', 'D\n2', 'Hy\n3', 'Pd\n4', 'Mf\n5', 'Pa\n6', 'Pt\n7', 'Sc\n8', 'Ma\n9', 'Si\n0']
    ext_scale = ['Mas', 'Dy', 'Do', 'Re', 'Cn']

    val_list = [
        pro_Point['L'],
        pro_Point['F'],
        pro_Point['K']
    ]
    cli_list = [
        pro_Point['Hs+0.5K'],
        pro_Point['D'],
        pro_Point['Hy'],
        pro_Point['Pd+0.4K'],
        pro_Point['Mf'],
        pro_Point['Pa'],
        pro_Point['Pt+1.0K'],
        pro_Point['Sc+1.0K'],
        pro_Point['Ma+0.2K'],
        pro_Point['Si']
    ]
    ext_list = [
        pro_Point['Mas'],
        pro_Point['Dy'],
        pro_Point['Do'],
        pro_Point['Re'],
        pro_Point['Cn']
    ]

    cli_max1 = max(cli_list)
    cli_max1_index = cli_list.index(cli_max1)
    if cli_max1_index != 9:
        first = cli_max1_index + 1
    else:
        first = 0
    cli_list[cli_max1_index] = 0
    cli_max2 = max(cli_list)
    cli_max2_index = cli_list.index(cli_max2)
    if cli_max2_index != 9:
        second = cli_max2_index + 1
    else:
        second = 0
    cli_list[cli_max1_index] = cli_max1
    two_point = '%s%s' % (str(first), str(second))

    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    fig = plt.figure(figsize=(10, 6), dpi=100, linewidth=1)
    ax = fig.add_subplot(111)
    ax.plot(range(len(val_list)), val_list, 'b*-')
    ax.plot(range(len(val_list), len(val_list) + len(cli_list)), cli_list, 'b*-')
    ax.plot(range(len(val_list) + len(cli_list), len(val_list) + len(cli_list) + len(ext_list)), ext_list, 'b*-')
    # plt.setp(ax.xaxis.get_majorticklabels(), rotation=-45)
    ax.set_xticks(range(len(val_list + cli_list + ext_list)))
    ax.set_xticklabels(val_scale + cli_scale + ext_scale)
    ax.set_yticks([0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 110, 120])
    ax.set_xlim(-0.5, len(val_scale + cli_scale + ext_scale) - 0.5)
    ax.set_ylim(0, 120)
    plt.axvline(2.5, ls="-", color="black")
    plt.axvline(12.5, ls="-", color="black")
    plt.axvline(7, ls="--", color="red")
    plt.axhline(50, ls="-", color="black")
    plt.axhline(60, ls="--", color="black")
    plt.axhline(70, ls="-", color="black")
    ax.plot(cli_max1_index + len(val_list), cli_max1, 'rp')
    ax.plot(cli_max2_index + len(val_list), cli_max2, 'rp')
    plt.annotate(r'$max1$', xy=(cli_max1_index + len(val_list)+0.1, cli_max1+2), color='red', fontsize=8)
    plt.annotate(r'$max2$', xy=(cli_max2_index + len(val_list)+0.1, cli_max2+2), color='red', fontsize=8)


def data_export():
    """
    导出数据
    export test data

    :return: None

    Note: Generate a '.xlsx' file to save the test information
    """
    print('请输入被试者姓名')
    name = input('> ')
    wb = Workbook()
    data_filename = time.strftime("%Y%m%d_%H%M_", time.localtime()) + name + '_MMPI测验'

    font1 = Font(name='黑体', size=12)
    font2 = Font(name='宋体', size=12)
    font3 = Font(name='Times New Roman', size=12, bold=True)
    font4 = Font(name='Times New Roman', size=12)
    alig1 = Alignment(horizontal='center', vertical='center')
    alig2 = Alignment(horizontal='general', vertical='center')

    # 表1，记录测验原始数据
    sheet1 = wb.active
    sheet1.title = '测验原始数据'
    sheet1['A1'] = '姓名'
    sheet1['A1'].font = font1
    sheet1['A1'].alignment = alig1
    sheet1['C1'] = '性别'
    sheet1['C1'].font = font1
    sheet1['C1'].alignment = alig1
    sheet1['E1'] = '年龄'
    sheet1['E1'].font = font1
    sheet1['E1'].alignment = alig1
    sheet1.merge_cells('A2:B2')
    sheet1['A2'] = '题 目'
    sheet1['A2'].font = font1
    sheet1['A2'].alignment = alig1
    sheet1.merge_cells('C2:D2')
    sheet1['C2'] = '回 答'
    sheet1['C2'].font = font1
    sheet1['C2'].alignment = alig1
    sheet1['B1'] = name
    sheet1['B1'].font = font2
    sheet1['B1'].alignment = alig1
    if Sex == '1':
        sex_name = '男'
    else:
        sex_name = '女'
    sheet1['D1'] = sex_name
    sheet1['D1'].font = font2
    sheet1['D1'].alignment = alig1
    sheet1['F1'] = Age
    sheet1['F1'].font = font4
    sheet1['F1'].alignment = alig1
    for i in range(len(Que)+1):
        if i == 73:
            if Sex == '1':
                temp_que = Que[i+1][Que[i+1].find('m')+1: Que[i+1].find('f')]
            else:
                temp_que = Que[i+1][Que[i+1].find('f') + 1:]
        elif i == len(Que):
            temp_que = '我保证是在专业人士指导下认真诚实地完成本次测验'
        else:
            temp_que = Que[i+1]

        sheet1['A%d' % (i + 3)].value = str(i + 1) + '.'
        sheet1['A%d' % (i + 3)].font = font2
        sheet1['A%d' % (i + 3)].alignment = alig1
        sheet1['B%d' % (i + 3)].value = temp_que
        sheet1['B%d' % (i + 3)].font = font2
        sheet1['B%d' % (i + 3)].alignment = alig2
        if Ans[i+1] == '1':
            temp_ans = '是'
            sheet1['C%d' % (i + 3)].value = temp_ans
            sheet1['C%d' % (i + 3)].font = font2
            sheet1['C%d' % (i + 3)].alignment = alig1
        else:
            temp_ans = '否'
            sheet1['D%d' % (i + 3)].value = temp_ans
            sheet1['D%d' % (i + 3)].font = font2
            sheet1['D%d' % (i + 3)].alignment = alig1

    # 表2，记录测验分数
    sheet2 = wb.create_sheet(title='测验分数')
    sheet2['A1'] = '姓名'
    sheet2['A1'].font = font1
    sheet2['A1'].alignment = alig1
    sheet2['C1'] = '性别'
    sheet2['C1'].font = font1
    sheet2['C1'].alignment = alig1
    sheet2['E1'] = '年龄'
    sheet2['E1'].font = font1
    sheet2['E1'].alignment = alig1
    sheet2['B1'].value = name
    sheet2['B1'].font = font2
    sheet2['B1'].alignment = alig1
    sheet2['D1'].value = sex_name
    sheet2['D1'].font = font2
    sheet2['D1'].alignment = alig1
    sheet2['F1'].value = Age
    sheet2['F1'].font = font4
    sheet2['F1'].alignment = alig1
    sheet2['A2'] = '量表类别'
    sheet2['A2'].font = font1
    sheet2['A2'].alignment = alig1
    sheet2['B2'] = '原始分'
    sheet2['B2'].font = font1
    sheet2['B2'].alignment = alig1
    sheet2['C2'] = '标准分（不加K）'
    sheet2['C2'].font = font1
    sheet2['C2'].alignment = alig1
    sheet2['D2'] = '标准分（加K）'
    sheet2['D2'].font = font1
    sheet2['D2'].alignment = alig1
    sheet2['A3'] = '*其中Q量表仅记录矛盾题的数量'
    sheet2['A3'].font = font2
    sheet2['A3'].alignment = alig2
    scale_list = ['Q*', 'L', 'F', 'K', 'Hs', 'D', 'Hy', 'Pd', 'Mf', 'Pa', 'Pt', 'Sc', 'Ma', 'Si',
                  'Mas', 'Dy', 'Do', 'Re', 'Cn']
    for i in range(len(scale_list)):
        sheet2['A%d' % (i+4)].value = scale_list[i]
        sheet2['A%d' % (i+4)].font = font3
        sheet2['A%d' % (i+4)].alignment = alig1
        sheet2['B%d' % (i+4)].value = ori_Point[scale_list[i]]
        sheet2['B%d' % (i+4)].font = font4
        sheet2['B%d' % (i+4)].alignment = alig1
        sheet2['C%d' % (i+4)].value = pro_Point[scale_list[i]]
        sheet2['C%d' % (i+4)].font = font4
        sheet2['C%d' % (i+4)].alignment = alig1
        if scale_list[i] == 'Hs':
            sheet2['D%d' % (i + 4)].value = pro_Point['Hs+0.5K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        elif scale_list[i] == 'Pd':
            sheet2['D%d' % (i + 4)].value = pro_Point['Pd+0.4K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        elif scale_list[i] == 'Pt':
            sheet2['D%d' % (i + 4)].value = pro_Point['Pt+1.0K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        elif scale_list[i] == 'Sc':
            sheet2['D%d' % (i + 4)].value = pro_Point['Sc+1.0K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        elif scale_list[i] == 'Ma':
            sheet2['D%d' % (i + 4)].value = pro_Point['Ma+0.2K']
            sheet2['D%d' % (i + 4)].font = font4
            sheet2['D%d' % (i + 4)].alignment = alig1
        else:
            pass
    sheet2['E2'] = '两点编码'
    sheet2['E2'].font = font1
    sheet2['E2'].alignment = alig1
    sheet2['F2'].value = two_point
    sheet2['F2'].font = font4
    sheet2['F2'].alignment = alig1

    wb.save(filename=data_filename + '.xlsx')
    plt.title('%s MMPI剖析图 加K分校正T分（中国常模）' % name)
    plt.savefig(data_filename)
